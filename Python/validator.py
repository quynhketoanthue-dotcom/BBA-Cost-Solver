# validator.py

import openpyxl
from pathlib import Path


class WorkbookValidator:
    """
    Validator v1

    Kiểm tra lại sheet Giá thành 2025.07 sau khi áp dụng:
    - Đọc cột C: Mã hàng
    - Đọc cột AK: Giá thành
    - Đọc cột AL: So sánh giá
    - Xuất danh sách mã còn âm
    """

    def __init__(self, workbook_file):
        self.workbook_file = Path(workbook_file)

    def validate_al(self):
        wb = openpyxl.load_workbook(self.workbook_file, data_only=True)
        ws = wb["Giá thành 2025.07"]

        negatives = []

        for r in range(12, ws.max_row + 1):
            ma_hang = ws[f"C{r}"].value
            al = ws[f"AL{r}"].value
            gia_thanh = ws[f"AK{r}"].value
            gia_ban = ws[f"D{r}"].value

            if not ma_hang:
                continue

            if isinstance(al, (int, float)) and al < 0:
                negatives.append({
                    "row": r,
                    "ma_hang": str(ma_hang).strip(),
                    "gia_ban": gia_ban,
                    "gia_thanh": gia_thanh,
                    "al": al,
                })

        return negatives

    def summary(self):
        negatives = self.validate_al()

        print()
        print("=" * 60)
        print("VALIDATOR SUMMARY")
        print("=" * 60)
        print("Số mã còn AL âm:", len(negatives))

        for item in negatives[:20]:
            print(
                item["ma_hang"],
                "AL=",
                round(item["al"], 2)
            )

        return negatives