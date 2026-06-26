import openpyxl

from models import Product


class WorkbookReader:

    def __init__(self, workbook):

        self.workbook = workbook

        self.wb = openpyxl.load_workbook(workbook, data_only=True)

        self.ws_cost = self.wb["Giá thành 2025.07"]
        self.ws_dm = self.wb["định mức"]
        self.ws_time = self.wb["Thời gian check"]
        self.ws_sale = self.wb["Sổ chi tiết bán hàng"]

    # ==========================
    # Định mức
    # ==========================

    def load_dinh_muc(self):

        data = {}

        for r in range(3, self.ws_dm.max_row + 1):

            ma = self.ws_dm[f"C{r}"].value

            if not ma:
                continue

            ma = str(ma).strip()

            data[ma] = {

                "dinh_muc": self.ws_dm[f"F{r}"].value,

                "ma_nvl": self.ws_dm[f"B{r}"].value

            }

        return data

    # ==========================
    # Thời gian
    # ==========================

    def load_time(self):

        data = {}

        for r in range(4, self.ws_time.max_row + 1):

            ma = self.ws_time[f"C{r}"].value

            if not ma:
                continue

            data[str(ma).strip()] = self.ws_time[f"D{r}"].value

        return data

    # ==========================
    # Bán hàng
    # ==========================

    def load_sale(self):

        data = {}

        for r in range(5, self.ws_sale.max_row + 1):

            ma = self.ws_sale[f"P{r}"].value

            if not ma:
                continue

            ma = str(ma).strip()

            tk_co = self.ws_sale[f"AO{r}"].value

            data[ma] = {

                "khach_hang": self.ws_sale[f"I{r}"].value,

                "ma_khach": self.ws_sale[f"H{r}"].value,

                "tk_co": tk_co,

                "khach_cap_nvl": str(tk_co).startswith("5113")

            }

        return data

    # ==========================
    # Đọc Product
    # ==========================

    def read_products(self):

        dm = self.load_dinh_muc()

        tg = self.load_time()

        sale = self.load_sale()

        products = []

        for r in range(12, self.ws_cost.max_row + 1):

            ma = self.ws_cost[f"C{r}"].value

            if not ma:
                continue

            ma = str(ma).strip()

            p = Product()

            p.row = r

            p.ma_hang = ma

            p.gia_ban = self.ws_cost[f"D{r}"].value

            p.gia_thanh = self.ws_cost[f"AK{r}"].value

            p.al = self.ws_cost[f"AL{r}"].value

            p.nhan_cong = self.ws_cost[f"AE{r}"].value

            # ----------------------

            if ma in dm:

                p.dinh_muc = dm[ma]["dinh_muc"]

                p.nvl = dm[ma]["ma_nvl"]

            # ----------------------

            if ma in tg:

                p.thoi_gian = tg[ma]

            # ----------------------

            if ma in sale:

                p.khach_hang = sale[ma]["khach_hang"]

                p.tk511 = sale[ma]["tk_co"]

                p.khach_cap_nvl = sale[ma]["khach_cap_nvl"]

            products.append(p)

        return products