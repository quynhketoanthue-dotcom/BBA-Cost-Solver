from pathlib import Path
import openpyxl

PROJECT = Path(__file__).resolve().parent.parent
INPUT = PROJECT / "Input"

file = list(INPUT.glob("*.xlsx"))[0]

wb = openpyxl.load_workbook(file, data_only=True)

for ws in wb.worksheets:

    print("=" * 80)
    print(ws.title)

    for r in range(1, 8):

        print(f"\n----- Dòng {r} -----")

        for c in range(1, min(ws.max_column, 40) + 1):

            v = ws.cell(r, c).value

            if v is not None:

                print(c, ":", v)