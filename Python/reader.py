import openpyxl

from models import Product, BomLine


class WorkbookReader:

    def __init__(self, workbook):
        self.workbook = workbook

        self.wb = openpyxl.load_workbook(workbook, data_only=True)

        self.ws_cost = self.wb["Giá thành 2025.07"]
        self.ws_dm = self.wb["định mức"]
        self.ws_time = self.wb["Thời gian check"]
        self.ws_sale = self.wb["Sổ chi tiết bán hàng"]
        self.ws_dvt = self.wb["ĐVT của NVL"]

    # ===========================
    # Helper
    # ===========================

    def _to_number(self, value, default=0):
        if isinstance(value, (int, float)):
            return float(value)

        try:
            return float(value)
        except Exception:
            return default

    def _clean(self, value):
        if value is None:
            return ""

        return str(value).strip()

    # ===========================
    # Đọc ĐVT NVL
    # ===========================

    def load_dvt(self):
        data = {}

        for r in range(3, self.ws_dvt.max_row + 1):
            ma_nvl = self.ws_dvt[f"B{r}"].value

            if not ma_nvl:
                continue

            ma_nvl = self._clean(ma_nvl)

            data[ma_nvl] = {
                "dvt": self.ws_dvt[f"D{r}"].value
            }

        return data

    # ===========================
    # Đọc BOM / Định mức
    # ===========================

    def load_bom(self):
        """
        Sheet định mức là BOM:
        1 mã hàng có thể có nhiều dòng NVL.

        Trả về:
        bom[ma_hang] = [BomLine, BomLine, ...]
        """

        dvt_map = self.load_dvt()

        bom = {}

        for r in range(3, self.ws_dm.max_row + 1):
            ma_nvl = self.ws_dm[f"B{r}"].value
            ma_hang = self.ws_dm[f"C{r}"].value
            dinh_muc = self.ws_dm[f"F{r}"].value

            if not ma_hang:
                continue

            ma_hang = self._clean(ma_hang)
            ma_nvl = self._clean(ma_nvl)

            line = BomLine()
            line.row = r
            line.ma_hang = ma_hang
            line.ma_nvl = ma_nvl
            line.ten_nvl = self.ws_dm[f"B{r}"].value
            line.dinh_muc = self._to_number(dinh_muc)

            if ma_nvl in dvt_map:
                line.dvt = dvt_map[ma_nvl].get("dvt", "")

            if ma_hang not in bom:
                bom[ma_hang] = []

            bom[ma_hang].append(line)

        return bom

    # ===========================
    # Đọc thời gian
    # ===========================

    def load_time(self):
        data = {}

        for r in range(4, self.ws_time.max_row + 1):
            ma_hang = self.ws_time[f"C{r}"].value

            if not ma_hang:
                continue

            ma_hang = self._clean(ma_hang)

            data[ma_hang] = {
                "row": r,
                "thoi_gian": self.ws_time[f"D{r}"].value
            }

        return data

    # ===========================
    # Đọc bán hàng
    # ===========================

    def load_sale(self):
        data = {}

        for r in range(5, self.ws_sale.max_row + 1):
            ma_hang = self.ws_sale[f"P{r}"].value

            if not ma_hang:
                continue

            ma_hang = self._clean(ma_hang)

            tk = self.ws_sale[f"AO{r}"].value

            # Nếu 1 mã hàng xuất hiện nhiều dòng, ưu tiên dòng có TK 5113 nếu có
            current = data.get(ma_hang)

            new_item = {
                "ma_khach": self.ws_sale[f"H{r}"].value,
                "khach_hang": self.ws_sale[f"I{r}"].value,
                "tk511": tk,
                "khach_cap_nvl": str(tk).startswith("5113")
            }

            if current is None:
                data[ma_hang] = new_item
            else:
                if new_item["khach_cap_nvl"]:
                    data[ma_hang] = new_item

        return data

    # ===========================
    # Đọc Product
    # ===========================

    def read_products(self):
        bom = self.load_bom()
        time_map = self.load_time()
        sale_map = self.load_sale()

        products = []

        for r in range(12, self.ws_cost.max_row + 1):
            ma_hang = self.ws_cost[f"C{r}"].value

            if not ma_hang:
                continue

            ma_hang = self._clean(ma_hang)

            p = Product()

            p.row = r
            p.ma_hang = ma_hang

            p.gia_ban = self.ws_cost[f"D{r}"].value
            p.gia_thanh = self.ws_cost[f"AK{r}"].value
            p.al = self.ws_cost[f"AL{r}"].value

            p.nhan_cong = self.ws_cost[f"AE{r}"].value

            # BOM
            if ma_hang in bom:
                p.bom_lines = bom[ma_hang]

                # Tổng định mức để tương thích optimizer/report hiện tại
                p.dinh_muc = sum(
                    self._to_number(line.dinh_muc)
                    for line in p.bom_lines
                )

                if p.bom_lines:
                    p.ma_nvl = p.bom_lines[0].ma_nvl
                    p.dvt = p.bom_lines[0].dvt

            # Thời gian
            if ma_hang in time_map:
                p.thoi_gian = time_map[ma_hang]["thoi_gian"]

            # Bán hàng / khách hàng / TK Có
            if ma_hang in sale_map:
                p.ma_khach = sale_map[ma_hang]["ma_khach"]
                p.khach_hang = sale_map[ma_hang]["khach_hang"]
                p.tk511 = sale_map[ma_hang]["tk511"]
                p.khach_cap_nvl = sale_map[ma_hang]["khach_cap_nvl"]

            products.append(p)

        return products